import openpyxl
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import sys

from twilio.rest import Client

import fbchat 
from getpass import getpass 
    
def whatsapp(name):
    
    # add path of chromedriver
    driver = webdriver.Chrome('D:\softwares\chromedriver')

    driver.get("https://web.whatsapp.com/")
    wait = WebDriverWait(driver, 600)

    # target is a person to whom message to be send
    target = name

    # birthday message
    string = "Heartly Felicitation on your Birthday "+name

    x_arg = '//span[contains(@title,' + target + ')]'
    group_title = wait.until(EC.presence_of_element_located((By.XPATH, x_arg)))
    group_title.click()


    message = driver.find_elements_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')[0]


    message.send_keys(string)

    sendbutton = driver.find_elements_by_xpath('//*[@id="main"]/footer/div[1]/div[3]/button')[0]
    sendbutton.click()

    driver.close()
    
    
def send_mail(rec, name):
    sender_email = "<email id >"
    rec_email = rec
    password = "<password>" #password
    subject = 'Birthday Wishes'
    message ='Heartly Felicitation on your Birthday'+name

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = rec_email
    msg['subject'] = subject

    msg.attach(MIMEText(message,'plain'))
    filename = "birth.gif"
    attachment = open(filename, 'rb')

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',"attachment; filename="+filename)
    msg.attach(part)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(sender_email, password)
    print("Login success")
    text = msg.as_string()
    server.sendmail(sender_email, rec_email, text) 
    print("Email has been sent to", rec_email)
    server.quit()
    
def sms(number):
    
    account_sid = '<account sid>'
    auth_token = '<auth token>' #auth tokenof your twilio account
    client = Client(account_sid, auth_token)

    message = client.messages         .create(
             body='Heartly Felicitation On Your Birthday..'+name,
             from_= '<your twilion number>' ,
             to= number
         )

    print("Birthday Message has sent successfully...")
    
# def facebook(name):
    
#     username = str("<username>") #enter login ID
#     client = fbchat.Client(username, getpass()) 
#     name = str(name) #friend's name who is having birthday
#     friends = client.searchForUsers(name)  # return a list of names 
#     friend = friends[0] 
#     msg = str("Heartly felicitation on your birthday"+name) 
#     sent = client.send(friend.uid, msg) 
#     if sent: 
# #         print("Message sent successfully!") 
        

today = date.today()
day = today.day
month = today.month


wb_obj = openpyxl.load_workbook("birthday.xlsx")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

for i in range(2, m_row+1):
    cell_obj = sheet_obj.cell(row = i, column = 3)
    a = cell_obj.value
    cell_obj = sheet_obj.cell(row = i, column = 4)
    b = cell_obj.value
    cell_obj = sheet_obj.cell(row = i, column = 2)
    rec = cell_obj.value
    cell_obj = sheet_obj.cell(row = i, column = 1)
    name = cell_obj.value
    cell_obj = sheet_obj.cell(row = i, column = 6)
    number = cell_obj.value
    
    if day == a and month == b:
        send_mail(rec, name)
        whatsapp(name)
        sms(number)
#         facebook(name)


# In[ ]:

